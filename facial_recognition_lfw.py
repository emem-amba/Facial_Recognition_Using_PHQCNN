"""
Facial Recognition using LFW Dataset and ResNet
Evaluates: Accuracy, Precision, Recall, F1-Score, and Computational Speed
"""
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import time
import warnings
from datetime import datetime

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, Dataset, Subset
import torchvision.transforms as transforms
import torchvision.models as models
from sklearn.datasets import fetch_lfw_people
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    confusion_matrix, classification_report, roc_curve, auc,
    precision_recall_curve
)
from sklearn.preprocessing import label_binarize

warnings.filterwarnings('ignore')

# Set random seeds for reproducibility
torch.manual_seed(42)
np.random.seed(42)

# Check for GPU availability
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f"Using device: {device}")

class LFWDataset(Dataset):
    """Custom Dataset for LFW faces"""
    def __init__(self, images, labels, transform=None):
        self.images = images
        self.labels = labels
        self.transform = transform
    
    def __len__(self):
        return len(self.labels)
    
    def __getitem__(self, idx):
        image = self.images[idx]
        # Convert to 3-channel image (grayscale to RGB)
        if len(image.shape) == 2:
            image = np.stack([image] * 3, axis=-1)
        
        # Normalize to [0, 1]
        image = image.astype(np.float32)
        if image.max() > 1:
            image = image / 255.0
        
        if self.transform:
            image = self.transform(image)
        
        return image, self.labels[idx]

def load_lfw_data(min_faces_per_person=70):
    """Load LFW dataset with specified minimum faces per person"""
    print("Loading LFW dataset...")
    start_time = time.time()
    
    lfw_people = fetch_lfw_people(
        min_faces_per_person=min_faces_per_person,
        resize=0.4,
        color=False
    )
    
    load_time = time.time() - start_time
    
    n_samples, h, w = lfw_people.images.shape
    X = lfw_people.images
    y = lfw_people.target
    target_names = lfw_people.target_names
    n_classes = len(target_names)
    
    print(f"Dataset loaded in {load_time:.2f} seconds")
    print(f"Total samples: {n_samples}")
    print(f"Image size: {h}x{w}")
    print(f"Number of classes: {n_classes}")
    print(f"Classes: {target_names}")
    
    return X, y, target_names, h, w, load_time

def create_resnet_model(num_classes, pretrained=True):
    """Create a modified ResNet18 for face recognition"""
    model = models.resnet18(weights='IMAGENET1K_V1' if pretrained else None)
    
    # Modify first conv layer to accept different input sizes
    model.conv1 = nn.Conv2d(3, 64, kernel_size=7, stride=2, padding=3, bias=False)
    
    # Modify final fully connected layer for our number of classes
    num_features = model.fc.in_features
    model.fc = nn.Sequential(
        nn.Dropout(0.5),
        nn.Linear(num_features, 256),
        nn.ReLU(),
        nn.Dropout(0.3),
        nn.Linear(256, num_classes)
    )
    
    return model

def train_model(model, train_loader, val_loader, criterion, optimizer, 
                num_epochs=20, device='cpu'):
    """Train the model and track metrics"""
    model = model.to(device)
    
    history = {
        'train_loss': [], 'val_loss': [],
        'train_acc': [], 'val_acc': [],
        'epoch_times': []
    }
    
    best_val_acc = 0.0
    best_model_state = None
    
    print("\nStarting training...")
    total_train_time = 0
    
    for epoch in range(num_epochs):
        epoch_start = time.time()
        
        # Training phase
        model.train()
        train_loss = 0.0
        train_correct = 0
        train_total = 0
        
        for images, labels in train_loader:
            images, labels = images.to(device), labels.to(device)
            
            optimizer.zero_grad()
            outputs = model(images)
            loss = criterion(outputs, labels)
            loss.backward()
            optimizer.step()
            
            train_loss += loss.item() * images.size(0)
            _, predicted = torch.max(outputs.data, 1)
            train_total += labels.size(0)
            train_correct += (predicted == labels).sum().item()
        
        train_loss = train_loss / train_total
        train_acc = train_correct / train_total
        
        # Validation phase
        model.eval()
        val_loss = 0.0
        val_correct = 0
        val_total = 0
        
        with torch.no_grad():
            for images, labels in val_loader:
                images, labels = images.to(device), labels.to(device)
                outputs = model(images)
                loss = criterion(outputs, labels)
                
                val_loss += loss.item() * images.size(0)
                _, predicted = torch.max(outputs.data, 1)
                val_total += labels.size(0)
                val_correct += (predicted == labels).sum().item()
        
        val_loss = val_loss / val_total
        val_acc = val_correct / val_total
        
        epoch_time = time.time() - epoch_start
        total_train_time += epoch_time
        
        # Save history
        history['train_loss'].append(train_loss)
        history['val_loss'].append(val_loss)
        history['train_acc'].append(train_acc)
        history['val_acc'].append(val_acc)
        history['epoch_times'].append(epoch_time)
        
        # Save best model
        if val_acc > best_val_acc:
            best_val_acc = val_acc
            best_model_state = model.state_dict().copy()
        
        print(f"Epoch [{epoch+1}/{num_epochs}] - "
              f"Train Loss: {train_loss:.4f}, Train Acc: {train_acc:.4f}, "
              f"Val Loss: {val_loss:.4f}, Val Acc: {val_acc:.4f}, "
              f"Time: {epoch_time:.2f}s")
    
    # Load best model
    if best_model_state is not None:
        model.load_state_dict(best_model_state)
    
    history['total_train_time'] = total_train_time
    print(f"\nTotal training time: {total_train_time:.2f} seconds")
    print(f"Best validation accuracy: {best_val_acc:.4f}")
    
    return model, history

def evaluate_model(model, test_loader, target_names, device='cpu'):
    """Comprehensive model evaluation"""
    model = model.to(device)
    model.eval()
    
    all_predictions = []
    all_labels = []
    all_probabilities = []
    
    inference_times = []
    
    print("\nEvaluating model...")
    
    with torch.no_grad():
        for images, labels in test_loader:
            images = images.to(device)
            
            # Measure inference time
            start_time = time.time()
            outputs = model(images)
            inference_time = time.time() - start_time
            inference_times.append(inference_time)
            
            probabilities = torch.softmax(outputs, dim=1)
            _, predicted = torch.max(outputs.data, 1)
            
            all_predictions.extend(predicted.cpu().numpy())
            all_labels.extend(labels.numpy())
            all_probabilities.extend(probabilities.cpu().numpy())
    
    all_predictions = np.array(all_predictions)
    all_labels = np.array(all_labels)
    all_probabilities = np.array(all_probabilities)
    
    # Calculate metrics
    metrics = {}
    
    # Overall metrics
    metrics['accuracy'] = accuracy_score(all_labels, all_predictions)
    metrics['precision_macro'] = precision_score(all_labels, all_predictions, average='macro', zero_division=0)
    metrics['precision_weighted'] = precision_score(all_labels, all_predictions, average='weighted', zero_division=0)
    metrics['recall_macro'] = recall_score(all_labels, all_predictions, average='macro', zero_division=0)
    metrics['recall_weighted'] = recall_score(all_labels, all_predictions, average='weighted', zero_division=0)
    metrics['f1_macro'] = f1_score(all_labels, all_predictions, average='macro', zero_division=0)
    metrics['f1_weighted'] = f1_score(all_labels, all_predictions, average='weighted', zero_division=0)
    
    # Per-class metrics
    precision_per_class = precision_score(all_labels, all_predictions, average=None, zero_division=0)
    recall_per_class = recall_score(all_labels, all_predictions, average=None, zero_division=0)
    f1_per_class = f1_score(all_labels, all_predictions, average=None, zero_division=0)
    
    metrics['precision_per_class'] = precision_per_class
    metrics['recall_per_class'] = recall_per_class
    metrics['f1_per_class'] = f1_per_class
    
    # Confusion matrix
    metrics['confusion_matrix'] = confusion_matrix(all_labels, all_predictions)
    
    # Computational metrics
    metrics['total_inference_time'] = sum(inference_times)
    metrics['avg_inference_time_per_batch'] = np.mean(inference_times)
    metrics['avg_inference_time_per_sample'] = sum(inference_times) / len(all_labels)
    metrics['throughput'] = len(all_labels) / sum(inference_times)  # samples per second
    
    # Classification report
    metrics['classification_report'] = classification_report(
        all_labels, all_predictions, 
        target_names=target_names,
        zero_division=0
    )
    
    # Store for ROC curves
    metrics['all_labels'] = all_labels
    metrics['all_probabilities'] = all_probabilities
    metrics['all_predictions'] = all_predictions
    
    return metrics

def create_metrics_dataframe(metrics, target_names, history, load_time):
    """Create comprehensive metrics DataFrames"""
    
    # Overall metrics DataFrame
    overall_data = {
        'Metric': [
            'Accuracy',
            'Precision (Macro)',
            'Precision (Weighted)',
            'Recall (Macro)',
            'Recall (Weighted)',
            'F1-Score (Macro)',
            'F1-Score (Weighted)',
            'Total Training Time (s)',
            'Total Inference Time (s)',
            'Avg Inference Time/Sample (ms)',
            'Throughput (samples/s)',
            'Data Load Time (s)'
        ],
        'Value': [
            f"{metrics['accuracy']:.4f}",
            f"{metrics['precision_macro']:.4f}",
            f"{metrics['precision_weighted']:.4f}",
            f"{metrics['recall_macro']:.4f}",
            f"{metrics['recall_weighted']:.4f}",
            f"{metrics['f1_macro']:.4f}",
            f"{metrics['f1_weighted']:.4f}",
            f"{history['total_train_time']:.2f}",
            f"{metrics['total_inference_time']:.4f}",
            f"{metrics['avg_inference_time_per_sample']*1000:.4f}",
            f"{metrics['throughput']:.2f}",
            f"{load_time:.2f}"
        ]
    }
    overall_df = pd.DataFrame(overall_data)
    
    # Per-class metrics DataFrame
    per_class_data = {
        'Class': target_names,
        'Precision': [f"{p:.4f}" for p in metrics['precision_per_class']],
        'Recall': [f"{r:.4f}" for r in metrics['recall_per_class']],
        'F1-Score': [f"{f:.4f}" for f in metrics['f1_per_class']]
    }
    per_class_df = pd.DataFrame(per_class_data)
    
    # Training history DataFrame
    history_data = {
        'Epoch': list(range(1, len(history['train_loss']) + 1)),
        'Train Loss': [f"{l:.4f}" for l in history['train_loss']],
        'Val Loss': [f"{l:.4f}" for l in history['val_loss']],
        'Train Accuracy': [f"{a:.4f}" for a in history['train_acc']],
        'Val Accuracy': [f"{a:.4f}" for a in history['val_acc']],
        'Epoch Time (s)': [f"{t:.2f}" for t in history['epoch_times']]
    }
    history_df = pd.DataFrame(history_data)
    
    return overall_df, per_class_df, history_df

def plot_all_visualizations(metrics, history, target_names, save_dir='/home/Final'):
    """Create all visualization plots"""

    # Ensure save directory exists
    os.makedirs(save_dir, exist_ok=True)

    
    # Set style
    plt.style.use('seaborn-v0_8-whitegrid')
    
    # 1. Training History Plot
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    
    epochs = range(1, len(history['train_loss']) + 1)
    
    # Loss plot
    axes[0].plot(epochs, history['train_loss'], 'b-o', label='Training Loss', markersize=4)
    axes[0].plot(epochs, history['val_loss'], 'r-o', label='Validation Loss', markersize=4)
    axes[0].set_xlabel('Epoch', fontsize=12)
    axes[0].set_ylabel('Loss', fontsize=12)
    axes[0].set_title('Training and Validation Loss', fontsize=14, fontweight='bold')
    axes[0].legend(fontsize=10)
    axes[0].grid(True, alpha=0.3)
    
    # Accuracy plot
    axes[1].plot(epochs, history['train_acc'], 'b-o', label='Training Accuracy', markersize=4)
    axes[1].plot(epochs, history['val_acc'], 'r-o', label='Validation Accuracy', markersize=4)
    axes[1].set_xlabel('Epoch', fontsize=12)
    axes[1].set_ylabel('Accuracy', fontsize=12)
    axes[1].set_title('Training and Validation Accuracy', fontsize=14, fontweight='bold')
    axes[1].legend(fontsize=10)
    axes[1].grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(f'{save_dir}/training_history.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved: training_history.png")
    
    # 2. Confusion Matrix
    fig, ax = plt.subplots(figsize=(10, 8))
    cm = metrics['confusion_matrix']
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', 
                xticklabels=target_names, yticklabels=target_names, ax=ax)
    ax.set_xlabel('Predicted Label', fontsize=12)
    ax.set_ylabel('True Label', fontsize=12)
    ax.set_title('Confusion Matrix', fontsize=14, fontweight='bold')
    plt.xticks(rotation=45, ha='right')
    plt.yticks(rotation=0)
    plt.tight_layout()
    plt.savefig(f'{save_dir}/confusion_matrix.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved: confusion_matrix.png")
    
    # 3. Per-class Metrics Bar Chart
    fig, ax = plt.subplots(figsize=(12, 6))
    x = np.arange(len(target_names))
    width = 0.25
    
    bars1 = ax.bar(x - width, metrics['precision_per_class'], width, label='Precision', color='#2ecc71')
    bars2 = ax.bar(x, metrics['recall_per_class'], width, label='Recall', color='#3498db')
    bars3 = ax.bar(x + width, metrics['f1_per_class'], width, label='F1-Score', color='#e74c3c')
    
    ax.set_xlabel('Class', fontsize=12)
    ax.set_ylabel('Score', fontsize=12)
    ax.set_title('Per-Class Performance Metrics', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(target_names, rotation=45, ha='right')
    ax.legend(fontsize=10)
    ax.set_ylim(0, 1.1)
    ax.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    plt.savefig(f'{save_dir}/per_class_metrics.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved: per_class_metrics.png")
    
    # 4. Overall Metrics Comparison
    fig, ax = plt.subplots(figsize=(10, 6))
    metrics_names = ['Accuracy', 'Precision\n(Macro)', 'Precision\n(Weighted)', 
                     'Recall\n(Macro)', 'Recall\n(Weighted)', 'F1-Score\n(Macro)', 'F1-Score\n(Weighted)']
    metrics_values = [
        metrics['accuracy'], metrics['precision_macro'], metrics['precision_weighted'],
        metrics['recall_macro'], metrics['recall_weighted'], 
        metrics['f1_macro'], metrics['f1_weighted']
    ]
    
    colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(metrics_names)))
    bars = ax.bar(metrics_names, metrics_values, color=colors, edgecolor='black', linewidth=1.2)
    
    # Add value labels on bars
    for bar, val in zip(bars, metrics_values):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01, 
                f'{val:.3f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
    
    ax.set_ylabel('Score', fontsize=12)
    ax.set_title('Overall Model Performance Metrics', fontsize=14, fontweight='bold')
    ax.set_ylim(0, 1.15)
    ax.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    plt.savefig(f'{save_dir}/overall_metrics.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved: overall_metrics.png")
    
    # 5. Computational Speed Analysis
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    
    # Epoch times
    axes[0].bar(range(1, len(history['epoch_times']) + 1), history['epoch_times'], 
                color='#9b59b6', edgecolor='black')
    axes[0].set_xlabel('Epoch', fontsize=12)
    axes[0].set_ylabel('Time (seconds)', fontsize=12)
    axes[0].set_title('Training Time per Epoch', fontsize=14, fontweight='bold')
    axes[0].grid(True, alpha=0.3, axis='y')
    
    # Speed metrics pie chart
    speed_labels = ['Data Loading', 'Training', 'Inference']
    speed_values = [load_time_global, history['total_train_time'], metrics['total_inference_time']]
    colors = ['#e74c3c', '#3498db', '#2ecc71']
    explode = (0.05, 0.05, 0.05)
    
    axes[1].pie(speed_values, labels=speed_labels, autopct='%1.1f%%', colors=colors,
                explode=explode, shadow=True, startangle=90)
    axes[1].set_title('Time Distribution', fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    plt.savefig(f'{save_dir}/computational_speed.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved: computational_speed.png")
    
    # 6. ROC Curves (One-vs-Rest for multi-class)
    n_classes = len(target_names)
    y_true_bin = label_binarize(metrics['all_labels'], classes=range(n_classes))
    y_score = metrics['all_probabilities']
    
    fig, ax = plt.subplots(figsize=(10, 8))
    colors = plt.cm.tab10(np.linspace(0, 1, n_classes))
    
    for i, (color, name) in enumerate(zip(colors, target_names)):
        fpr, tpr, _ = roc_curve(y_true_bin[:, i], y_score[:, i])
        roc_auc = auc(fpr, tpr)
        ax.plot(fpr, tpr, color=color, lw=2, label=f'{name} (AUC = {roc_auc:.3f})')
    
    ax.plot([0, 1], [0, 1], 'k--', lw=2, label='Random Classifier')
    ax.set_xlabel('False Positive Rate', fontsize=12)
    ax.set_ylabel('True Positive Rate', fontsize=12)
    ax.set_title('ROC Curves (One-vs-Rest)', fontsize=14, fontweight='bold')
    ax.legend(loc='lower right', fontsize=9)
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(f'{save_dir}/roc_curves.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved: roc_curves.png")
    
    # 7. Summary Dashboard
    fig = plt.figure(figsize=(16, 12))
    
    # Create grid
    gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3)
    
    # Training curves
    ax1 = fig.add_subplot(gs[0, 0])
    ax1.plot(epochs, history['train_acc'], 'b-', label='Train', linewidth=2)
    ax1.plot(epochs, history['val_acc'], 'r-', label='Val', linewidth=2)
    ax1.set_title('Accuracy Over Epochs', fontweight='bold')
    ax1.set_xlabel('Epoch')
    ax1.set_ylabel('Accuracy')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Loss curves
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.plot(epochs, history['train_loss'], 'b-', label='Train', linewidth=2)
    ax2.plot(epochs, history['val_loss'], 'r-', label='Val', linewidth=2)
    ax2.set_title('Loss Over Epochs', fontweight='bold')
    ax2.set_xlabel('Epoch')
    ax2.set_ylabel('Loss')
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    # Metrics summary
    ax3 = fig.add_subplot(gs[0, 2])
    metric_labels = ['Accuracy', 'Precision', 'Recall', 'F1-Score']
    metric_vals = [metrics['accuracy'], metrics['precision_weighted'], 
                   metrics['recall_weighted'], metrics['f1_weighted']]
    bars = ax3.barh(metric_labels, metric_vals, color=['#3498db', '#2ecc71', '#e74c3c', '#9b59b6'])
    ax3.set_xlim(0, 1)
    ax3.set_title('Overall Metrics (Weighted)', fontweight='bold')
    for bar, val in zip(bars, metric_vals):
        ax3.text(val + 0.02, bar.get_y() + bar.get_height()/2, f'{val:.3f}', va='center')
    ax3.grid(True, alpha=0.3, axis='x')
    
    # Confusion matrix
    ax4 = fig.add_subplot(gs[1, :2])
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', 
                xticklabels=target_names, yticklabels=target_names, ax=ax4)
    ax4.set_title('Confusion Matrix', fontweight='bold')
    ax4.set_xlabel('Predicted')
    ax4.set_ylabel('True')
    
    # Per-class F1
    ax5 = fig.add_subplot(gs[1, 2])
    y_pos = np.arange(len(target_names))
    ax5.barh(y_pos, metrics['f1_per_class'], color='#e74c3c')
    ax5.set_yticks(y_pos)
    ax5.set_yticklabels(target_names)
    ax5.set_xlabel('F1-Score')
    ax5.set_title('Per-Class F1-Score', fontweight='bold')
    ax5.grid(True, alpha=0.3, axis='x')
    
    # Time distribution
    ax6 = fig.add_subplot(gs[2, 0])
    time_labels = ['Load', 'Train', 'Inference']
    time_vals = [load_time_global, history['total_train_time'], metrics['total_inference_time']]
    ax6.bar(time_labels, time_vals, color=['#e74c3c', '#3498db', '#2ecc71'])
    ax6.set_ylabel('Time (seconds)')
    ax6.set_title('Time Distribution', fontweight='bold')
    ax6.grid(True, alpha=0.3, axis='y')
    
    # Speed metrics text
    ax7 = fig.add_subplot(gs[2, 1:])
    ax7.axis('off')
    speed_text = f"""
    COMPUTATIONAL PERFORMANCE SUMMARY
    ═══════════════════════════════════════════
    
    Data Loading Time:           {load_time_global:.2f} seconds
    Total Training Time:         {history['total_train_time']:.2f} seconds
    Average Epoch Time:          {np.mean(history['epoch_times']):.2f} seconds
    Total Inference Time:        {metrics['total_inference_time']:.4f} seconds
    Inference Time per Sample:   {metrics['avg_inference_time_per_sample']*1000:.4f} ms
    Throughput:                  {metrics['throughput']:.2f} samples/second
    
    ═══════════════════════════════════════════
    Model: ResNet-18 (Modified for Face Recognition)
    Dataset: Labeled Faces in the Wild (LFW)
    Device: {device}
    """
    ax7.text(0.1, 0.5, speed_text, transform=ax7.transAxes, fontsize=11,
             verticalalignment='center', fontfamily='monospace',
             bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
    
    plt.suptitle('Facial Recognition with ResNet on LFW Dataset - Complete Analysis', 
                 fontsize=16, fontweight='bold', y=1.02)
    
    plt.savefig(f'{save_dir}/summary_dashboard.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved: summary_dashboard.png")
    
    print("\nAll visualizations saved successfully!")

def save_results_to_excel(overall_df, per_class_df, history_df, metrics, save_path):
    """Save all results to an Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Overall Metrics
    ws1 = wb.active
    ws1.title = 'Overall Metrics'
    
    ws1['A1'] = 'FACIAL RECOGNITION EVALUATION METRICS'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:B1')
    
    ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1['A3'] = f'Model: ResNet-18'
    ws1['A4'] = f'Dataset: LFW (Labeled Faces in the Wild)'
    
    start_row = 6
    for r_idx, row in enumerate(dataframe_to_rows(overall_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20
    
    # Sheet 2: Per-Class Metrics
    ws2 = wb.create_sheet('Per-Class Metrics')
    
    ws2['A1'] = 'PER-CLASS PERFORMANCE METRICS'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:D1')
    
    start_row = 3
    for r_idx, row in enumerate(dataframe_to_rows(per_class_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
    
    for col in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col].width = 20
    
    # Sheet 3: Training History
    ws3 = wb.create_sheet('Training History')
    
    ws3['A1'] = 'TRAINING HISTORY'
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A1:F1')
    
    start_row = 3
    for r_idx, row in enumerate(dataframe_to_rows(history_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 18
    
    # Sheet 4: Confusion Matrix
    ws4 = wb.create_sheet('Confusion Matrix')
    
    ws4['A1'] = 'CONFUSION MATRIX'
    ws4['A1'].font = Font(bold=True, size=14)
    
    cm = metrics['confusion_matrix']
    target_names = per_class_df['Class'].tolist()
    
    # Header row
    ws4.cell(row=3, column=1, value='Predicted →')
    for j, name in enumerate(target_names):
        cell = ws4.cell(row=3, column=j+2, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Data rows
    for i, name in enumerate(target_names):
        cell = ws4.cell(row=i+4, column=1, value=name)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='D9E2F3')
        cell.alignment = center_align
        
        for j in range(len(target_names)):
            cell = ws4.cell(row=i+4, column=j+2, value=int(cm[i, j]))
            cell.border = thin_border
            cell.alignment = center_align
    
    wb.save(save_path)
    print(f"\nResults saved to: {save_path}")

# Main execution
if __name__ == "__main__":
    print("="*60)
    print("FACIAL RECOGNITION USING LFW DATASET AND RESNET")
    print("="*60)
    
    # Load data
    X, y, target_names, h, w, load_time = load_lfw_data(min_faces_per_person=70)
    load_time_global = load_time
    
    # Split data
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.25, stratify=y, random_state=42
    )
    X_train, X_val, y_train, y_val = train_test_split(
        X_train, y_train, test_size=0.15, stratify=y_train, random_state=42
    )
    
    print(f"\nData split:")
    print(f"  Training samples: {len(X_train)}")
    print(f"  Validation samples: {len(X_val)}")
    print(f"  Test samples: {len(X_test)}")
    
    # Define transforms
    transform = transforms.Compose([
        transforms.ToTensor(),
        transforms.Resize((128, 128)),
        transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])
    ])
    
    # Create datasets
    train_dataset = LFWDataset(X_train, y_train, transform=transform)
    val_dataset = LFWDataset(X_val, y_val, transform=transform)
    test_dataset = LFWDataset(X_test, y_test, transform=transform)
    
    # Create data loaders
    batch_size = 32
    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True, num_workers=0)
    val_loader = DataLoader(val_dataset, batch_size=batch_size, shuffle=False, num_workers=0)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=False, num_workers=0)
    
    # Create model
    num_classes = len(target_names)
    model = create_resnet_model(num_classes, pretrained=True)
    
    # Count parameters
    total_params = sum(p.numel() for p in model.parameters())
    trainable_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    print(f"\nModel parameters:")
    print(f"  Total: {total_params:,}")
    print(f"  Trainable: {trainable_params:,}")
    
    # Define loss and optimizer
    criterion = nn.CrossEntropyLoss()
    optimizer = optim.Adam(model.parameters(), lr=0.001, weight_decay=1e-4)
    
    # Train model
    model, history = train_model(
        model, train_loader, val_loader, criterion, optimizer,
        num_epochs=15, device=device
    )
    
    # Evaluate model
    metrics = evaluate_model(model, test_loader, target_names, device=device)
    
    # Print results
    print("\n" + "="*60)
    print("EVALUATION RESULTS")
    print("="*60)
    print(f"\nOverall Metrics:")
    print(f"  Accuracy:            {metrics['accuracy']:.4f}")
    print(f"  Precision (Macro):   {metrics['precision_macro']:.4f}")
    print(f"  Precision (Weighted):{metrics['precision_weighted']:.4f}")
    print(f"  Recall (Macro):      {metrics['recall_macro']:.4f}")
    print(f"  Recall (Weighted):   {metrics['recall_weighted']:.4f}")
    print(f"  F1-Score (Macro):    {metrics['f1_macro']:.4f}")
    print(f"  F1-Score (Weighted): {metrics['f1_weighted']:.4f}")
    
    print(f"\nComputational Speed:")
    print(f"  Total Training Time:       {history['total_train_time']:.2f} seconds")
    print(f"  Avg Epoch Time:            {np.mean(history['epoch_times']):.2f} seconds")
    print(f"  Total Inference Time:      {metrics['total_inference_time']:.4f} seconds")
    print(f"  Inference Time per Sample: {metrics['avg_inference_time_per_sample']*1000:.4f} ms")
    print(f"  Throughput:                {metrics['throughput']:.2f} samples/second")
    
    print("\nClassification Report:")
    print(metrics['classification_report'])
    
    # Create DataFrames
    overall_df, per_class_df, history_df = create_metrics_dataframe(
        metrics, target_names, history, load_time
    )
    
    # Save to Excel
    save_results_to_excel(
        overall_df, per_class_df, history_df, metrics,
        '/home/facial_recognition_results.xlsx'
    )
    
    # Create visualizations
    plot_all_visualizations(metrics, history, target_names, '/home')
    
    print("\n" + "="*60)
    print("PROCESS COMPLETE")
    print("="*60)

